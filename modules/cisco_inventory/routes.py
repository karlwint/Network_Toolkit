"""
Cisco Switch Stack Inventory — Blueprint
Web-based port of cisco_inventory_gui.py.

Stack-aware inventory collection running:
  show version, show switch detail, show inventory,
  show interfaces status, show cdp neighbors

Outputs a 4-sheet Excel workbook:
  Summary, Stack Overview, Stack Members, Hardware Inventory
"""

import re
import os
import threading
import io
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Blueprint, render_template, jsonify, request, session, send_file
from modules import get_cisco_credentials, require_cisco_creds

inventory_bp = Blueprint('inventory', __name__, template_folder='../../templates')

_inv_jobs = {}
_inv_lock = threading.Lock()


# ══════════════════════════════════════════════════════════════
#  PARSERS (from cisco_inventory_gui.py)
# ══════════════════════════════════════════════════════════════

def _parse_show_switch_detail(raw):
    """Parse 'show switch detail' — handles 9K and 3750/3650 formats."""
    members = []

    # Format A (9K): two tables — version table then role table
    sw_model_map = {}
    in_ver_table = False
    for line in raw.splitlines():
        if re.search(r'Model\s+SW Version', line, re.IGNORECASE):
            in_ver_table = True
            continue
        if in_ver_table:
            m = re.match(r'^\s*\*?(\d+)\s+(\d+)\s+(\S+)\s+(\S+)\s+(\S+)', line)
            if m:
                sw_model_map[m.group(1)] = {
                    "ports": m.group(2), "model": m.group(3),
                    "sw_version": m.group(4), "sw_image": m.group(5),
                }
            elif line.strip() == "" and sw_model_map:
                in_ver_table = False

    in_role_table = False
    for line in raw.splitlines():
        if re.search(r'Role\s+Mac Address\s+Priority', line, re.IGNORECASE):
            in_role_table = True
            continue
        if in_role_table:
            m = re.match(
                r'^\s*\*?(\d+)\s+(\S+)\s+([\da-fA-F]{4}\.[\da-fA-F]{4}\.[\da-fA-F]{4})\s+(\d+)\s+(\S+)\s+(\S+)',
                line)
            if m:
                num = m.group(1)
                extra = sw_model_map.get(num, {})
                members.append({
                    "member": num, "role": m.group(2), "mac": m.group(3),
                    "priority": m.group(4), "hw_ver": m.group(5), "state": m.group(6),
                    "model": extra.get("model", ""), "ports": extra.get("ports", ""),
                    "sw_version": extra.get("sw_version", ""), "sw_image": extra.get("sw_image", ""),
                })
            elif line.strip() == "" and members:
                break

    if members:
        return members

    # Format B (3750/3650)
    in_b = False
    for line in raw.splitlines():
        if re.search(r'Role\s+Priority\s+State', line, re.IGNORECASE):
            in_b = True
            continue
        if in_b:
            m = re.match(
                r'^\s*\*?(\d+)\s+(\S+)\s+(\d+)\s+(\S+)\s+\S+\s+([\da-fA-F]{4}\.[\da-fA-F]{4}\.[\da-fA-F]{4})',
                line)
            if m:
                members.append({
                    "member": m.group(1), "role": m.group(2), "priority": m.group(3),
                    "state": m.group(4), "mac": m.group(5),
                    "hw_ver": "", "model": "", "ports": "", "sw_version": "", "sw_image": "",
                })
            elif line.strip() == "" and members:
                break

    if members:
        return members

    # Fallback: single switch
    m = re.search(r'\*?1\s+\S+\s+([\da-fA-F]{4}\.[\da-fA-F]{4}\.[\da-fA-F]{4})', raw)
    return [{
        "member": "1", "role": "Active", "mac": m.group(1) if m else "",
        "priority": "", "hw_ver": "", "state": "Ready",
        "model": "", "ports": "", "sw_version": "", "sw_image": "",
    }]


def _parse_show_inventory(raw):
    """Parse 'show inventory' — tag each component with stack member number."""
    components = []
    stanzas = re.split(r'(?=^NAME\s*:)', raw, flags=re.MULTILINE)
    for stanza in stanzas:
        stanza = stanza.strip()
        if not stanza:
            continue
        name_m = re.search(r'NAME\s*:\s*"([^"]*)"', stanza)
        descr_m = re.search(r'DESCR\s*:\s*"([^"]*)"', stanza)
        pid_m = re.search(r'PID\s*:\s*(\S+)', stanza)
        vid_m = re.search(r'VID\s*:\s*(\S+)', stanza)
        sn_m = re.search(r'SN\s*:\s*(\S+)', stanza)
        if not (name_m or pid_m or sn_m):
            continue
        name = name_m.group(1).strip() if name_m else ""
        descr = descr_m.group(1).strip() if descr_m else ""
        pid = pid_m.group(1).strip() if pid_m else ""
        vid = vid_m.group(1).strip() if vid_m else ""
        sn = sn_m.group(1).strip() if sn_m else ""

        stack_member = ""
        m = re.match(r'[Ss]witch\s+(\d+)', name)
        if m:
            stack_member = m.group(1)
        if not stack_member:
            m = re.match(r'[A-Za-z]{1,4}(\d+)/\d+/\d+', name)
            if m:
                stack_member = m.group(1)
        if not stack_member:
            m = re.search(r'\b(\d+)/\d+\b', name)
            if m:
                stack_member = m.group(1)

        components.append({
            "stack_member": stack_member, "name": name, "descr": descr,
            "pid": pid, "vid": vid, "sn": sn,
        })
    return components


def _parse_show_version(ver):
    d = {"hostname": "", "model": "", "ios_version": "", "serial": "",
         "uptime": "", "ios_image": "", "total_ram": "", "flash": ""}
    m = re.search(r"^(\S+)\s+uptime is (.+)$", ver, re.MULTILINE)
    if m:
        d["hostname"] = m.group(1)
        d["uptime"] = m.group(2).strip()
    m = re.search(r"Cisco IOS.*?Version\s+([\S]+)", ver)
    if m:
        d["ios_version"] = m.group(1).rstrip(",")
    m = re.search(r"System image file is \"(.+?)\"", ver)
    if m:
        d["ios_image"] = m.group(1)
    for pat in [r"Model [Nn]umber\s*:\s*(\S+)",
                r"cisco\s+(WS-\S+|C\d{4}\S*|ISR\S*|ASR\S*)",
                r"Cisco\s+(Catalyst\s+\S+)"]:
        m = re.search(pat, ver)
        if m:
            d["model"] = m.group(1).strip()
            break
    m = re.search(r"[Ss]ystem [Ss]erial [Nn]umber\s*:\s*(\S+)", ver)
    if m:
        d["serial"] = m.group(1)
    else:
        m = re.search(r"Processor board ID\s+(\S+)", ver)
        if m:
            d["serial"] = m.group(1)
    m = re.search(r"(\d+)K bytes of physical memory", ver)
    if m:
        d["total_ram"] = m.group(1) + "K"
    m = re.search(r"(\d+)K bytes of.*?[Ff]lash", ver)
    if m:
        d["flash"] = m.group(1) + "K"
    return d


def _parse_interfaces(intf):
    conn_c = notconn_c = dis_c = other_c = 0
    for line in intf.splitlines():
        if re.match(r"^(Gi|Fa|Te|Hu|Tw|Fo|Et|Po)\S+\s", line):
            lo = line.lower()
            if "connected" in lo and "notconnect" not in lo:
                conn_c += 1
            elif "notconnect" in lo:
                notconn_c += 1
            elif "disabled" in lo:
                dis_c += 1
            else:
                other_c += 1
    return {"total": conn_c + notconn_c + dis_c + other_c,
            "connected": conn_c, "notconnect": notconn_c, "disabled": dis_c}


# ══════════════════════════════════════════════════════════════
#  SSH COLLECTION
# ══════════════════════════════════════════════════════════════

def _collect_one(ip, label, username, password, enable):
    """Collect all data from one stack IP. Returns (stack_row, member_rows, hw_rows)."""
    from netmiko import ConnectHandler, NetmikoTimeoutException, NetmikoAuthenticationException

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    stack_row = {
        "label": label, "ip": ip, "status": "Failed",
        "hostname": "", "model": "", "ios_version": "", "serial": "",
        "uptime": "", "ios_image": "", "total_ram": "", "flash": "",
        "total": "", "connected": "", "notconnect": "", "disabled": "",
        "cdp_neighbors": "", "mgmt_vlan": "",
        "stack_size": 0, "hw_component_count": 0, "collected_at": ts,
    }
    member_rows = []
    hw_rows = []

    params = {
        "device_type": "cisco_ios", "host": ip,
        "username": username, "password": password,
        "secret": enable or password,
        "timeout": 45, "conn_timeout": 15, "banner_timeout": 15,
    }
    try:
        with ConnectHandler(**params) as conn:
            conn.enable()
            hostname_raw = conn.find_prompt().replace("#", "").replace(">", "").strip()

            ver = conn.send_command("show version", use_textfsm=False)
            sw = conn.send_command("show switch detail", use_textfsm=False)
            inv = conn.send_command("show inventory", use_textfsm=False)
            intf = conn.send_command("show interfaces status", use_textfsm=False)
            cdp = conn.send_command("show cdp neighbors", use_textfsm=False)
            vlan = conn.send_command(
                "show run | include management-vlan|ip default-gateway",
                use_textfsm=False)

            vd = _parse_show_version(ver)
            stack_row.update(vd)
            if not stack_row["hostname"]:
                stack_row["hostname"] = hostname_raw

            stack_row.update(_parse_interfaces(intf))

            neighbors = set(re.findall(r"^(\S+)\s", cdp, re.MULTILINE))
            neighbors.discard("Device")
            stack_row["cdp_neighbors"] = len(neighbors)

            m = re.search(r"management-vlan\s+(\d+)", vlan)
            if m:
                stack_row["mgmt_vlan"] = m.group(1)

            members = _parse_show_switch_detail(sw)
            stack_row["stack_size"] = len(members)
            hostname = stack_row["hostname"] or label

            for mbr in members:
                member_rows.append({
                    "ip": ip, "label": label, "hostname": hostname,
                    "member": mbr["member"], "role": mbr["role"],
                    "state": mbr["state"], "priority": mbr["priority"],
                    "mac": mbr["mac"], "hw_ver": mbr["hw_ver"],
                    "model": mbr["model"], "ports": mbr["ports"],
                    "sw_version": mbr["sw_version"], "sw_image": mbr["sw_image"],
                    "collected_at": ts,
                })

            components = _parse_show_inventory(inv)
            stack_row["hw_component_count"] = len(components)
            mbr_model = {m["member"]: m["model"] for m in members}

            for comp in components:
                mem_num = comp["stack_member"]
                hw_rows.append({
                    "ip": ip, "label": label, "hostname": hostname,
                    "stack_member": mem_num,
                    "member_model": mbr_model.get(mem_num, "") if mem_num else "",
                    "name": comp["name"], "descr": comp["descr"],
                    "pid": comp["pid"], "vid": comp["vid"], "sn": comp["sn"],
                    "collected_at": ts,
                })

            stack_row["status"] = "OK"

    except NetmikoAuthenticationException:
        stack_row["status"] = "Auth Failed"
    except NetmikoTimeoutException:
        stack_row["status"] = "Timeout"
    except Exception as e:
        stack_row["status"] = f"Error: {str(e)[:80]}"

    return stack_row, member_rows, hw_rows


# ══════════════════════════════════════════════════════════════
#  EXCEL WRITER
# ══════════════════════════════════════════════════════════════

STACK_COLS = [
    ("Label / Name", "label", 22), ("IP Address", "ip", 16),
    ("Status", "status", 12), ("Hostname", "hostname", 20),
    ("Active Model", "model", 20), ("IOS Version", "ios_version", 20),
    ("Active Serial", "serial", 20), ("Stack Size", "stack_size", 11),
    ("Uptime", "uptime", 32), ("IOS Image", "ios_image", 42),
    ("RAM", "total_ram", 12), ("Flash", "flash", 12),
    ("Total Ports", "total", 12), ("Connected", "connected", 12),
    ("Not Connected", "notconnect", 14), ("Disabled", "disabled", 10),
    ("CDP Neighbors", "cdp_neighbors", 15), ("Mgmt VLAN", "mgmt_vlan", 12),
    ("HW Components", "hw_component_count", 15), ("Collected At", "collected_at", 20),
]

MEMBER_COLS = [
    ("IP Address", "ip", 16), ("Label / Name", "label", 22),
    ("Hostname", "hostname", 20), ("Member #", "member", 10),
    ("Role", "role", 12), ("State", "state", 12),
    ("Priority", "priority", 10), ("MAC Address", "mac", 18),
    ("Model", "model", 20), ("Ports", "ports", 8),
    ("SW Version", "sw_version", 16), ("SW Image", "sw_image", 30),
    ("HW Version", "hw_ver", 12), ("Collected At", "collected_at", 20),
]

HW_COLS = [
    ("IP Address", "ip", 16), ("Label / Name", "label", 22),
    ("Hostname", "hostname", 20), ("Stack Member", "stack_member", 13),
    ("Member Model", "member_model", 20), ("Component Name", "name", 32),
    ("Description", "descr", 52), ("PID", "pid", 22),
    ("VID", "vid", 10), ("Serial Number", "sn", 22),
    ("Collected At", "collected_at", 20),
]


def _write_excel(stack_rows, member_rows, hw_rows):
    """Generate Excel workbook in memory, return BytesIO."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    F_HDR_BLUE = PatternFill("solid", start_color="1F4E79")
    F_HDR_GREEN = PatternFill("solid", start_color="1A4731")
    F_HDR_PURP = PatternFill("solid", start_color="3D1A6E")
    F_ALT_BLUE = PatternFill("solid", start_color="EBF3FB")
    F_ALT_GREEN = PatternFill("solid", start_color="EBF7F1")
    F_ALT_PURP = PatternFill("solid", start_color="F3EEFB")
    F_OK = PatternFill("solid", start_color="C6EFCE")
    F_WARN = PatternFill("solid", start_color="FFEB9C")
    F_FAIL = PatternFill("solid", start_color="FFC7CE")
    F_ACTIVE = PatternFill("solid", start_color="D6EAF8")
    F_CHASSIS = PatternFill("solid", start_color="D6E4F0")
    SIDE = Side(style="thin", color="BFBFBF")
    BORD = Border(left=SIDE, right=SIDE, top=SIDE, bottom=SIDE)

    def hdr_row(ws, cols, fill):
        for ci, (h, _, w) in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            c.fill = fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORD
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

    wb = Workbook()

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Cisco Switch Stack Inventory Report"
    ws["A1"].font = Font(bold=True, size=16, name="Arial", color="1F4E79")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(italic=True, size=10, name="Arial", color="595959")

    ok_stacks = sum(1 for r in stack_rows if r["status"] == "OK")
    active_members = sum(1 for m in member_rows if m.get("role", "").lower() in ("active", "master"))
    unique_pids = len({h["pid"] for h in hw_rows if h["pid"]})

    stats = [
        ("Total Stacks Polled", len(stack_rows)),
        ("Successfully Polled", ok_stacks),
        ("Failed", len(stack_rows) - ok_stacks),
        ("Total Physical Switches", len(member_rows)),
        ("Active/Master Members", active_members),
        ("Total HW Components", len(hw_rows)),
        ("Unique PIDs", unique_pids),
    ]
    for i, (lbl, val) in enumerate(stats, start=4):
        ws[f"A{i}"] = lbl; ws[f"A{i}"].font = Font(bold=True, name="Arial", size=11)
        ws[f"B{i}"] = val; ws[f"B{i}"].font = Font(name="Arial", size=11)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12

    # Stack Overview
    wo = wb.create_sheet("Stack Overview")
    hdr_row(wo, STACK_COLS, F_HDR_BLUE)
    for ri, row in enumerate(stack_rows, 2):
        st = row.get("status", "")
        sf = F_OK if st == "OK" else (F_WARN if st in ("Auth Failed", "Timeout") else F_FAIL)
        alt = F_ALT_BLUE if ri % 2 == 0 else None
        for ci, (_, key, _) in enumerate(STACK_COLS, 1):
            c = wo.cell(row=ri, column=ci, value=row.get(key, ""))
            c.font = Font(name="Arial", size=10)
            c.border = BORD
            c.alignment = Alignment(vertical="center")
            if key == "status":
                c.fill = sf
                c.font = Font(bold=True, name="Arial", size=10)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif alt:
                c.fill = alt
    wo.auto_filter.ref = wo.dimensions

    # Stack Members
    wm = wb.create_sheet("Stack Members")
    hdr_row(wm, MEMBER_COLS, F_HDR_PURP)
    for ri, row in enumerate(member_rows, 2):
        is_active = row.get("role", "").lower() in ("active", "master")
        alt = F_ACTIVE if is_active else (F_ALT_PURP if ri % 2 == 0 else None)
        for ci, (_, key, _) in enumerate(MEMBER_COLS, 1):
            c = wm.cell(row=ri, column=ci, value=row.get(key, ""))
            c.font = Font(name="Arial", size=10, bold=is_active)
            c.border = BORD
            c.alignment = Alignment(vertical="center")
            if key == "role" and is_active:
                c.font = Font(bold=True, name="Arial", size=10, color="1F4E79")
            if alt:
                c.fill = alt
    wm.auto_filter.ref = wm.dimensions

    # Hardware Inventory
    wh = wb.create_sheet("Hardware Inventory")
    hdr_row(wh, HW_COLS, F_HDR_GREEN)
    for ri, row in enumerate(hw_rows, 2):
        is_sw_root = bool(re.match(r'^[Ss]witch\s+\d+$', row["name"].strip()))
        is_chassis = any(kw in row["name"].lower()
                         for kw in ("chassis", "supervisor", "switch system", "base board", "motherboard"))
        if is_sw_root:
            fill = F_CHASSIS
        elif is_chassis:
            fill = PatternFill("solid", start_color="D6EAF8")
        else:
            fill = F_ALT_GREEN if ri % 2 == 0 else None

        for ci, (_, key, _) in enumerate(HW_COLS, 1):
            c = wh.cell(row=ri, column=ci, value=row.get(key, ""))
            c.font = Font(name="Arial", size=10, bold=(is_sw_root or is_chassis))
            c.border = BORD
            c.alignment = Alignment(vertical="center", wrap_text=(key == "descr"))
            if fill:
                c.fill = fill
    wh.auto_filter.ref = wh.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════
#  ASYNC JOB
# ══════════════════════════════════════════════════════════════

class InventoryJob:
    def __init__(self, job_id):
        self.id = job_id
        self.status = "pending"
        self.progress = 0
        self.total = 0
        self.stack_rows = []
        self.member_rows = []
        self.hw_rows = []
        self.log = []
        self.stats = {"stacks": 0, "ok": 0, "failed": 0, "members": 0, "hw_components": 0}
        self.cancel_event = threading.Event()

    def to_dict(self):
        return {
            "id": self.id, "status": self.status,
            "progress": self.progress, "total": self.total,
            "log": self.log[-100:], "stats": self.stats,
            "stack_rows": self.stack_rows,
            "member_rows": self.member_rows[-50:],
        }


# ─── Routes ───────────────────────────────────────────────────

@inventory_bp.route('/')
def inventory_index():
    creds = get_cisco_credentials()
    return render_template('inventory/index.html',
                           has_creds=bool(creds['username'] and creds['password']))


@inventory_bp.route('/api/scan', methods=['POST'])
@require_cisco_creds
def start_inventory():
    """Start stack inventory collection."""
    data = request.json
    entries_raw = data.get('entries', [])  # List of "ip" or "ip | label" strings
    threads = min(50, max(1, data.get('threads', 8)))

    # Parse entries: "10.0.0.1" or "10.0.0.1 | Core Switch"
    entries = []
    for line in entries_raw:
        line = line.strip()
        if not line:
            continue
        if '|' in line:
            parts = line.split('|', 1)
            entries.append((parts[0].strip(), parts[1].strip()))
        else:
            entries.append((line, line))

    if not entries:
        return jsonify({"error": "No IP addresses provided"}), 400

    job_id = "inv_" + datetime.now().strftime('%Y%m%d_%H%M%S')
    job = InventoryJob(job_id)
    job.total = len(entries)
    job.stats["stacks"] = len(entries)
    job.status = "running"

    with _inv_lock:
        _inv_jobs[job_id] = job

    # Capture credentials NOW while we're in request context
    creds = get_cisco_credentials()

    def run_inventory():
        job.log.append(f"[START] {len(entries)} stack IP(s) • {threads} threads")

        with ThreadPoolExecutor(max_workers=threads) as pool:
            fmap = {
                pool.submit(_collect_one, ip, lbl,
                            creds['username'], creds['password'], creds['enable']): (ip, lbl)
                for ip, lbl in entries
            }
            for future in as_completed(fmap):
                if job.cancel_event.is_set():
                    break
                sr, mrs, hws = future.result()
                job.stack_rows.append(sr)
                job.member_rows.extend(mrs)
                job.hw_rows.extend(hws)
                job.progress += 1

                if sr["status"] == "OK":
                    job.stats["ok"] += 1
                else:
                    job.stats["failed"] += 1
                job.stats["members"] = len(job.member_rows)
                job.stats["hw_components"] = len(job.hw_rows)

                status = "OK" if sr["status"] == "OK" else "FAIL"
                job.log.append(
                    f"  [{status:4s}] {sr['ip']:18s} {sr.get('hostname') or sr['label']:22s} "
                    f"stack:{sr.get('stack_size', 0)} members  {len(hws)} HW components")

        if job.cancel_event.is_set():
            job.status = "cancelled"
            job.log.append("\n[CANCELLED] Stopped by user.")
        else:
            job.status = "complete"
            job.log.append(
                f"\n[COMPLETE] {job.stats['ok']} stacks polled • "
                f"{len(job.member_rows)} physical switches • "
                f"{len(job.hw_rows)} hardware components")

    thread = threading.Thread(target=run_inventory, daemon=True)
    thread.start()
    return jsonify({"job_id": job_id, "total": len(entries)})


@inventory_bp.route('/api/job/<job_id>')
def get_inventory_job(job_id):
    with _inv_lock:
        job = _inv_jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify(job.to_dict())


@inventory_bp.route('/api/job/<job_id>/cancel', methods=['POST'])
def cancel_inventory_job(job_id):
    with _inv_lock:
        job = _inv_jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    job.cancel_event.set()
    return jsonify({"success": True})


@inventory_bp.route('/api/job/<job_id>/export', methods=['GET'])
def export_inventory(job_id):
    """Export inventory results as a 4-sheet Excel workbook."""
    with _inv_lock:
        job = _inv_jobs.get(job_id)
    if not job or not job.stack_rows:
        return jsonify({"error": "No results to export"}), 404

    buf = _write_excel(job.stack_rows, job.member_rows, job.hw_rows)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'cisco_stack_inventory_{ts}.xlsx')
